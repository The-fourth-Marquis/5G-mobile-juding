# coding='utf-8'
import os
import time
import tkinter as tk
from tkinter import ttk, messagebox
import requests
import json
import openpyxl
from openpyxl import Workbook, load_workbook
import unicodedata
from collections import OrderedDict
import jieba
import jieba.analyse
import numpy as np
from PIL import Image, ImageTk
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from snownlp import SnowNLP
import os
import dashscope
import pypandoc
from http import HTTPStatus
from urllib.parse import quote
from tkinter.scrolledtext import ScrolledText
import threading
from queue import Queue

dashscope.api_key = "sk-c1f5a9a769314fdfa2d35d8aa0966854"
stop = False
queue = Queue()
Product_id = 100000000000


#展示生成的文件
def open_output_folder():
    """
    打开/output文件夹

    该函数检查指定的`output_folder_path`是否存在。
    存在则使用`os.startfile()`打开文件夹；否则打印消息告知用户文件夹不存在。
    """
    output_folder_path = f"{Product_id}"
    if os.path.exists(output_folder_path):
        os.startfile(output_folder_path)
    else:
        messagebox.showerror("错误", "输出文件夹不存在")   

def load_and_display_images(directory_path, master):
    """从指定目录加载并显示所有PNG图像"""
    images = [f for f in os.listdir(directory_path) if f.endswith('.png')]
    row = 3
    column = 0
    for image_name in images:
        image_path = os.path.join(directory_path, image_name)

        original_image = Image.open(image_path)
        
        # 计算等比例缩放尺寸
        width, height = original_image.size
        scale_factor = 300 / float(height)
        new_width = int(width * scale_factor)

        resized_image = original_image.resize((new_width, 300),Image.Resampling.LANCZOS)
        photo_image = ImageTk.PhotoImage(resized_image)
        
        label = tk.Label(master, image=photo_image)
        label.image = photo_image  # 防止图片被垃圾回收
        label.grid(row=row, column=column, padx=10, pady=10)
        column += 1

def list_files_in_directory(directory_path):
    """列出指定目录下的所有文件名"""
    files = [f for f in os.listdir(directory_path) if os.path.isfile(os.path.join(directory_path, f))]
    return files

def update_listbox(directory_path, listbox, self):
    """根据输入的目录路径，更新列表框显示的文件名"""
    files = list_files_in_directory(directory_path)
    listbox.delete(0, tk.END)  # 清空当前列表
    for file in files:
        listbox.insert(tk.END, file)

    load_and_display_images(directory_path, self)

#展示评论
def update_text_area(self,text_widget):
    """从队列中获取内容并更新Text部件"""
    while not queue.empty():
        content = queue.get()
        text_widget.config(state=tk.NORMAL)
        text_widget.insert(tk.END, content)
        text_widget.see(tk.END)  # 自动滚动到底部
        text_widget.config(state=tk.DISABLED)
    self.after(100, update_text_area,self, text_widget)  # 每100毫秒检查一次队列

def on_another_done(message):
    """another函数完成后处理逻辑,如显示错误消息等"""
    if message:
        messagebox.showerror("错误", message)

def worker(items, j, product_id):
    """后台线程执行的任务"""
    global stop
    try:
        if not os.path.exists(f"{product_id}"):
            os.makedirs(f"{product_id}")
        new_table = f'{product_id}/original_data.xlsx'
        index = (j - 1) * 10 + 2  

        data = openpyxl.load_workbook(new_table)
        ws = data.active

        it = True
        for test in items:
            for i in range(4):
                queue.put(test[i])  # 将内容放入队列
                ws.cell(row=index, column=i + 1).value = test[i]
            queue.put('_______________________\n')  # 分割线也放入队列
            if it:
                it = False
            index += 1

        data.save(new_table)

        if it and j == 3:
            stop = True
            on_another_done("评论数量不足")

    except Exception as e:
        on_another_done(f"发生错误: {e}")

def another(items, j, product_id):
    """包装worker函数,确保在新线程中运行"""
    thread = threading.Thread(target=worker, args=(items, j, product_id))
    thread.start()

#配置与调用通义千问API
# 1. 定义生成商品链接的函数
def generate_good_link(product_id):
    """
    根据给定的商品ID,生成京东商品链接。

    参数:
    - product_id (int): 商品ID

    返回:
    str: 生成的商品链接字符串
    """
    base_url = "https://item.jd.com/{pid}.html"
    return base_url.format(pid=product_id)

# 2. 读取TXT文件内容的函数
def read_txt_file(file_path):
    """
    读取指定路径下的TXT文件内容。如果读取过程中发生错误,打印错误信息并返回None。

    参数:
    - file_path (str): TXT文件路径

    返回:
    str or None: 成功读取则返回文件内容,否则返回None
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            txt_content = f.read()
        return txt_content
    except Exception as e:
        messagebox.showerror("错误", f"读取TXT文件时发生错误: {e}")
        return None

# 3. Markdown转Docx的函数
def markdown_to_docx(content, output_path):
    """
    将Markdown格式的内容转换为Docx文档,并保存到指定路径。转换过程中可能发生的错误会被捕获并打印。

    参数:
    - content (str): Markdown格式的内容
    - output_path (str): 目标Docx文档的保存路径
    """
    # 将Markdown内容暂存到临时文件
    with open('temp.md', 'w', encoding='utf-8') as f:
        f.write(content)

    try:
        converted_text = pypandoc.convert_file('temp.md', 'docx', outputfile=output_path)
    except (IOError, OSError, RuntimeError) as e:
        messagebox.showerror("错误",f"转换过程中发生错误: {e}")
    else:
        messagebox.showinfo("提示","------ 分析报告已生成并保存至{} ------".format(output_path))

    # 删除临时文件
    os.remove('temp.md')

# 4. 构建系统消息的函数
def build_system_message(product_id):
    """
    根据商品ID构建一条系统消息,包含商品链接和提示用户撰写分析报告的文本。

    参数:
    - product_id (int): 商品ID

    返回:
    dict: 系统消息字典
    """
    good_link = generate_good_link(product_id)
    return {
        'role': 'system',
        'content': f'针对京东商品({quote(good_link)}),我将提供商品链接及用户评价摘要,基于这些信息,请撰写一份简洁的商品分析报告,并给出具有针对性的改进建议。'
    }

# 5. 主要调用逻辑的函数
def call_with_messages(product_id):
    """
    根据商品ID,执行一系列操作以生成商品分析报告并保存为Docx文档。

    参数:
    - product_id (int): 商品ID
    """
    with open(f"{product_id}/contents.txt", 'r', encoding='utf-8') as file:
        txt_summary = file.read()[:5500]

    good_link = generate_good_link(product_id)

    system_message = build_system_message(product_id)
    user_message = {
        'role': 'user',
        'content': [{'text': f"{good_link}{txt_summary}"}]
    }

    messages = [system_message, user_message]

    response = dashscope.Generation.call(
        dashscope.Generation.Models.qwen_turbo,
        messages=messages,
        result_format='message',
    )

    if response.status_code == HTTPStatus.OK:
        recipe = response['output']['choices'][0]['message']['content']

        # 将Markdown格式的结果转换为Word文档
        markdown_to_docx(recipe, f"{product_id}/report.docx")
    else:
        print('请求失败：',
              f"Request id: {response.request_id}, ",
              f"Status code: {response.status_code}, ",
              f"Error code: {response.code}, ",
              f"Error message: {response.message}")

# 6. 主函数
def qianwen(product_id):
    """
    解析命令行参数,获取商品ID,并调用call_with_messages()函数生成分析报告。
    """
    call_with_messages(product_id)

#存入Excel
def read_excel(product_id):
    """
    主程序: 将Excel文件数据转为易于阅读的txt
    """
    workbook = openpyxl.load_workbook(f'{product_id}/cleaned_data.xlsx')
    target_column_number = 3
    with open(f'{product_id}/contents.txt', 'w', encoding='utf-8') as f:
        for row in workbook.active.iter_rows():
            content = row[target_column_number - 1].value
            f.write(content + '\n')

    f.close()

    #messagebox.showinfo("提示",'评论数据已总结')

# 自然语言处理
def nlp(product_id):
    # 1. 打开并读取已分词的文本文件
    source = open(f"{product_id}/segmented_words.txt", "r", encoding='utf8')
    line = source.readlines()

    # 2. 对读取的每一行文本进行情感分析,并将结果存入sentimentslist列表
    sentimentslist = []
    for i in line:
        s = SnowNLP(i)
        sentimentslist.append(s.sentiments)

    # 3. 统计情感得分中积极和消极的数量
    positive = len([s for s in sentimentslist if s >= 0.5])
    negative = len([s for s in sentimentslist if s < 0.5])

    # 4. 定义饼图的标签、大小及颜色
    labels = ['Positive', 'Negative']
    sizes = [positive, negative]
    colors = ['green', 'red']

    # 5. 绘制饼图并设置样式
    plt.pie(sizes,
            labels=labels,
            colors=colors,
            autopct='%1.1f%%')
    plt.title('Analysis of Sentiments')
    plt.savefig(f"{product_id}/sentiment_analysis_pie_chart.png",
                dpi=300,
                bbox_inches='tight')

#文本分词处理
def chinese_word_segmentation(file_path, product_id):
    """
    分词并生成词云

    参数:
    - file_path (str): Excel文件路径

    步骤：
    1. 从Excel文件中提取评论内容进行分词
    2. 过滤停用词并保存分词结果到文本文件
    3. 统计分词结果的词频
    4. 使用背景图片和词频数据生成词云
    5. 保存词云图片
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 设置停用词
    jieba.analyse.set_stop_words("data/stopwords.txt")
    stop_words = set()
    with open('data/stopwords.txt', 'r', encoding='utf-8') as f:
        for word in f:
            stop_words.add(word.strip())

    # 从Excel文件中提取评论内容并分词
    segmented_words = []
    for row_index in range(2, sheet.max_row + 1):
        content = sheet.cell(row=row_index, column=3).value
        if content is not None:
            words = jieba.cut(content)
            words_filtered = ' '.join(word for word in words if word not in stop_words)
            segmented_words.append(words_filtered)

    # 保存分词结果到文本文件
    with open(f'{product_id}/segmented_words.txt', 'w', encoding='utf-8') as file:
        for words in segmented_words:
            file.write(words + '\n')
    print("Chinese word segmentation is completed and saved in 'segmented_words.txt'.")

    # 统计词频
    word_freq = {}
    for words in segmented_words:
        for word in words.split():
            if word not in stop_words:
                if word not in word_freq:
                    word_freq[word] = 0
                word_freq[word] += 1

    # 加载背景图片并生成词云
    img = Image.open("data/bg.png")
    mask = np.array(img)

    wordcloud = WordCloud(width=1000, height=1000, mask=mask, background_color='white', font_path='STKAITI.TTF',
                          stopwords=stop_words, random_state=50, max_words=40)
    wordcloud.generate_from_frequencies(word_freq)

    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')
    wordcloud.to_file(f'{product_id}/wordcloud.png')

def segmented(product_id):
    """
    主程序：获取商品评论数据,并将数据写入Excel文件
    """
    chinese_word_segmentation(f'{product_id}/cleaned_data.xlsx',product_id)

#数据清洗
def remove_duplicates_and_emojis(original_file, new_file):
    print("开始处理原始文件: ", original_file)

    wb = openpyxl.load_workbook(original_file)
    ws = wb.active

    # 读取表头
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active

    # 将表头复制到新工作簿中
    for i, header in enumerate(headers, start=1):
        ws_new.cell(row=1, column=i).value = header

    # 使用有序字典存储去重后的评论及其对应列数据
    comments_dict = OrderedDict()

    # 遍历原始数据,去除表情符号并去重
    for row in range(2, ws.max_row + 1):
        content = ws.cell(row=row, column=3).value
        if not content:
            continue

        content_no_emoji = remove_emojis(content)
        if content_no_emoji not in comments_dict:
            comments_dict[content_no_emoji] = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]

    # 将去重后的内容写入新工作簿
    for row_index, content in enumerate(comments_dict.values(), start=2):
        for col_index, cell_value in enumerate(content):
            if col_index == 2:  # 对第3列（即评论列）单独去除表情符号
                cell_value = remove_emojis(cell_value)
            ws_new.cell(row=row_index, column=col_index + 1).value = cell_value

    wb_new.save(new_file)

    print("数据清理完毕,表情已删除。结果保存至文件: ", new_file)

def remove_emojis(text):
    """
    移除文本中的表情符号

    参数:
    - text (str): 输入文本

    返回:
    - str: 去除表情符号后的文本
    """
    clean_text = ""
    for char in text:
        if unicodedata.category(char) != 'So':
            clean_text += char
    return clean_text

def clean(product_id):
    remove_duplicates_and_emojis(f'{product_id}/original_data.xlsx', f'{product_id}/cleaned_data.xlsx')

#爬虫操作
def start(page, product_id, score):
    """
    发起请求并获取指定商品评分、页码的评论数据

    参数:
    - page (int): 页码
    - product_id (int): 商品ID
    - score (int): 评论类型,4为全部评论

    返回:
    - dict: 解析后的JSON数据
    """
    url = f'https://club.jd.com/comment/productPageComments.action?&productId={product_id}&score={score}&sortType=5&page={page}&pageSize=10&isShadowSku=0&fold=1'

    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Mobile Safari/537.36"
    }
    time.sleep(2)
    response = requests.get(url=url, headers=headers)
    data = json.loads(response.text)
    return data

def parse(data):
    """
    解析评论数据,生成包含用户昵称、评论ID、评论内容、创建时间的元组序列

    参数:
    - data (dict): JSON格式的评论数据

    返回:
    - generator: 包含用户昵称、评论ID、评论内容、创建时间的元组序列
    """
    items = data['comments']
    for i in items:
        yield (
            i['nickname'],
            i['id'],
            i['content'],
            i['creationTime']
        )

def excel(items, product_id):
    """
    将评论数据写入Excel文件(如不存在则创建)

    参数:
    - items (generator): 包含用户昵称、评论ID、评论内容、创建时间的元组序列
    - product_id (int): 商品ID
    """
    if not os.path.exists(f"{product_id}"):
        os.makedirs(f"{product_id}")
    new_table = f'{product_id}/original_data.xlsx'
    wb = Workbook()
    ws = wb.active

    # 设置表头
    head_data = ['nickname', 'id', '内容', '时间']
    for i in range(0, 4):
        ws.cell(row=1, column=i + 1).value = head_data[i]

    index = 2

    # 写入评论数据
    for data in items:
        for i in range(0, 4):
            print(data[i])
            ws.cell(row=index, column=i + 1).value = data[i]
        print('______________________')
        index += 1

    wb.save(new_table)


def spider(product_id):
    """
    主程序：获取商品评论数据,并将数据写入Excel文件
    """
    global stop
    stop = False
    
    score = 4
    page_amount = 20

    j = 1
    judge = True

    for i in range(0, page_amount):
        time.sleep(1.5)
        first = start(j, product_id, score)
        
        test = parse(first)

        if judge:
            excel(test, product_id)
            judge = False
        else:
            another(test, j, product_id)

        if stop:
            return

        print(f'第{j}页抓取完毕')
        time.sleep(0.1)
        j += 1

# 处理输入输出
def update_progress(root, progress, progress_bar):
    """
    更新进度条值并刷新界面

    参数:
    - progress (int): 进度值(0-100)
    - progress_bar (ttk.Progressbar): 要更新的进度条对象
    """
    progress_bar["value"] = progress
    root.update_idletasks()

def simulate_progress(root, progress_bar):
    """
    模拟进度更新,用于演示进度条功能

    参数:
    - progress_bar (ttk.Progressbar): 要模拟更新的进度条对象
    """
    for _ in range(100):
        update_progress(root, progress_bar["value"] + 1, progress_bar)
        time.sleep(0.05)

def run_commands(root, product_id):
    """
    根据商品ID执行一系列命令并更新进度条

    参数:
    - product_id (str): 商品ID
    """
    global progress_bar
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode='determinate')
    progress_bar.grid(row=4, column=1, columnspan=6, sticky=tk.EW, pady=10)
    progress_bar.start()
   
    spider(product_id)
    if not stop:
        update_progress(root, 10, progress_bar)
        clean(product_id)
        if not stop:
            update_progress(root, 10, progress_bar)
            segmented(product_id)
            if not stop:
                update_progress(root, 10, progress_bar)
                nlp(product_id)
                if not stop:
                    update_progress(root, 10, progress_bar)
                    read_excel(product_id)
                    if not stop:
                        update_progress(root, 10, progress_bar)
                        qianwen(product_id)
                        if not stop:
                            update_progress(root, 10, progress_bar)
                            messagebox.showinfo("提示", "报告已生成完成！")

    progress_bar.stop()
    
def on_click(root, entry):
    """
    处理"开始分析"按钮点击事件

    获取输入框中的商品ID,如果非空,则调用`run_commands()`；否则显示错误消息。
    """
    product_id = entry.get()
    global Product_id
    Product_id = product_id
    if product_id:
        run_commands(root, product_id)
    else:
        messagebox.showerror("错误", "请先输入商品ID")

class MultiPageApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # 初始化主窗口
        self.geometry("1200x700")
        self.resizable(True, True)
        self.title("5G手机评价分析")
        
        # 创建一个容器来容纳所有页面
        self.container = tk.Frame(self)
        self.container.pack(side="top", fill="both", expand=True)
        
        # 配置容器的排版管理器，以便页面可以填满容器
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)
        
        # 页面字典，用于存储页面及其对应的Frame
        self.frames = {}
        
        # 动态添加页面
        for F in (PageOne, PageTwo):
            page_name = F.__name__
            frame = F(parent=self.container, controller=self)
            self.frames[page_name] = frame
            
            # 将Frame添加到容器中，但初始时都不可见
            frame.grid(row=0, column=0, sticky="nsew")

        # 默认显示第一个页面
        self.show_frame("PageOne")

    def show_frame(self, page_name):
        """显示指定的页面"""
        frame = self.frames[page_name]
        frame.tkraise()
    
# 定义各个页面类
class PageOne(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        ttk.Label(self, text="请输入京东商品ID:", font=("Arial", 14)).grid(row=0, column=0, padx=10, pady=10)
        entry = ttk.Entry(self, width=25)
        entry.grid(row=0, column=3, padx=10)

        start_button = ttk.Button(self, text="开始分析", command=lambda: on_click(self, entry), width=15)
        start_button.grid(row=0, column=5, pady=10)

        button = ttk.Button(self, text="进入数据页",
                           command=lambda: controller.show_frame("PageTwo"), width=15)
        button.grid(row=1, column=5, pady=10)

        label = ttk.Label(self, text="评论数据", font=("Helvetica", 12))
        label.grid(row=1, column=1, pady=1)

        text_area = ScrolledText(self, state='disabled', height=20)
        text_area.grid(row=2, column=1,columnspan=20, pady=1)

        update_text_area(self,text_area)  # 启动UI更新循环

class PageTwo(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = ttk.Label(self, text=f"当前的商品ID为{Product_id},爬取以及生成的数据如下:", font=("Helvetica", 16))
        label.grid(row=0, column=0, pady=10, sticky= "W")  # 修改为grid布局
        
        button = ttk.Button(self, text="返回爬取页",
                           command=lambda: controller.show_frame("PageOne"))
        button.grid(row=0, column=10, pady=10)

         # 列表框用于显示文件名
        listbox = tk.Listbox(self, width=25, height=10)
        listbox.grid(row=1, column=0, pady=10, sticky= "W")

        # 按钮点击事件，用于更新列表框内容
        browse_button = ttk.Button(self, text="浏览", command=lambda: update_listbox(f"{Product_id}", listbox, self))
        browse_button.grid(row=0, column=5, pady=10)

        output_button = ttk.Button(self, text=" 查看文件详情", command=open_output_folder, width=15)
        output_button.grid(row=0, column=20, pady=10)

        label = ttk.Label(self, text="对当前的商品用户反馈的关注倾向词云以及情感倾向饼图:", font=("Helvetica", 16))
        label.grid(row=2, column=0, pady=10, sticky= "W")
   

if __name__ == "__main__":
    app = MultiPageApp()
    app.mainloop()
